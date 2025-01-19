import os
import datetime
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from projects.models import Project

class Command(BaseCommand):
    help = "Import projects from an Excel file"
    file_path = './Erfolge Digitale Transformation_Digital Lab.xlsx'


    def handle(self, *args, **kwargs):
        file_path = './Erfolge Digitale Transformation_Digital Lab.xlsx'
        file_path = './Kurzbeschriebe_Projekte-Smart City-Fundus.xlsx'
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File '{file_path}' does not exist."))
            return

        self.stdout.write(f"Loading data from {file_path}...")
        if file_path == './Erfolge Digitale Transformation_Digital Lab.xlsx':
            try:
                workbook = load_workbook(filename=file_path)
                sheet = workbook.active

                # Assuming the first row contains headers
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    title, description, contact_person, link, strategie_erwaehnt, beitrag, umsetzung, portfoliomanagement = row
                    if title:
                        project = Project(
                            title=title,
                            decription=description,
                            url = link,
                            contact_person=contact_person or 'NA',
                            start_year=2024,
                            end_year=2025,
                            duration_months=12,
                            progress=umsetzung or 0,
                            verwendung_portal=True,
                        )
                        project.save()

                self.stdout.write(self.style.SUCCESS("Projects imported successfully."))

            except Exception as e:
                self.stdout.write(self.style.ERROR(f"Error occurred: {str(e)}"))
            else:
                try:
                    workbook = load_workbook(filename=file_path)
                    sheet = workbook.active
                    
                    # Assuming the first row contains headers
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        title, beschreibung, federführung, link = row
                        print(row)
                        if title:
                            print(title)
                            project = Project(
                                title=title,
                                decription=beschreibung,
                                url = link,
                                contact_person=federführung or 'NA',
                                start_year=2024,
                                end_year=2025,
                                duration_months=12
                            )
                            project.save()

                    self.stdout.write(self.style.SUCCESS("Projects imported successfully."))

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"Error occurred: {str(e)}"))